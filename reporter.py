"""
reporter.py — Генерація Excel-звіту для AutoPhotoSorter.

Створює файл .xlsx з двома аркушами:
  1. «Зведений звіт»  — по одному рядку на папку
  2. «Деталі файлів»  — по одному рядку на кожне зображення
"""

import os
from datetime import datetime
import logging

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from analyzer import CATEGORY_LABELS_UK

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Стилі
# ---------------------------------------------------------------------------

_DARK_BLUE = "1F4E79"
_LIGHT_BLUE = "BDD7EE"
_GREEN = "E2EFDA"
_YELLOW = "FFF2CC"
_RED = "FCE4D6"
_GRAY = "F2F2F2"
_WHITE = "FFFFFF"

_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_FONT = Font(name="Calibri", bold=True, color=_WHITE, size=11)
_HEADER_FILL = PatternFill(start_color=_DARK_BLUE, end_color=_DARK_BLUE, fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_CELL_ALIGN_LEFT = Alignment(vertical="center", wrap_text=True)
_CELL_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _style_header_row(ws: Worksheet, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER
    ws.row_dimensions[row].height = 36


def _apply_col_widths(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Аркуш 1 — Зведений звіт
# ---------------------------------------------------------------------------

_SUMMARY_HEADERS = [
    "№",
    "Папка",
    "Статус",
    "Кількість фото",
    "Головне фото",
    "Метод аналізу",
    "Перейменовано",
    "Примітки",
]

_SUMMARY_WIDTHS = [5, 32, 22, 16, 38, 18, 16, 65]


def _build_summary_sheet(ws: Worksheet, results: list[dict], title_row: int = 2) -> None:
    # Назва таблиці (рядок 1)
    # Об'єднуємо перші N-1 стовпців під заголовок, останній залишаємо для мітки часу
    merge_end_col = len(_SUMMARY_HEADERS) - 1
    ws.merge_cells(f"A1:{get_column_letter(merge_end_col)}1")
    title_cell = ws["A1"]
    title_cell.value = "AutoPhotoSorter — Звіт сортування фотографій товарів"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=_DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Мітка часу в останньому стовпці рядка 1 (поза межами merge)
    ts_cell = ws.cell(row=1, column=len(_SUMMARY_HEADERS))
    ts_cell.value = f"Згенеровано: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ts_cell.font = Font(name="Calibri", size=9, color="666666")
    ts_cell.alignment = Alignment(horizontal="right", vertical="center")

    # Заголовки
    for col, header in enumerate(_SUMMARY_HEADERS, start=1):
        ws.cell(row=title_row, column=col, value=header)
    _style_header_row(ws, title_row, len(_SUMMARY_HEADERS))

    # Дані
    for row_offset, result in enumerate(results):
        row = title_row + 1 + row_offset
        folder_name = result.get("folder_name", "")
        error = result.get("error")

        # Статус і заливка рядка
        if error == "No images found":
            status = "⚠️ Немає фото"
            row_fill = _fill(_YELLOW)
        elif error:
            status = "❌ Помилка"
            row_fill = _fill(_RED)
        elif result.get("fallback_used"):
            status = "⚠️ Потрібна перевірка"
            row_fill = _fill(_YELLOW)
        else:
            status = "✅ Успішно"
            row_fill = _fill(_GREEN)

        # Головне фото
        if result.get("has_ideal_main"):
            main_status = "✅ Ідеальне знайдено"
        elif result.get("fallback_used") and result.get("fallback_image"):
            fb = result["fallback_image"]
            main_status = f"⚠️ Альтернатива: {fb.get('filename', '')}"
        else:
            main_status = "❌ Не визначено"

        # Методи аналізу
        methods = {img.get("method", "") for img in result.get("sorted_images", [])}
        methods.discard("")
        method_str = ", ".join(sorted(methods)) if methods else "—"

        # Примітки
        notes = []
        if result.get("fallback_used"):
            notes.append(
                f"Папка [{folder_name}]: Ідеальне головне фото не знайдено. "
                "Обрано альтернативу. Потрібно знайти/зробити нове головне фото."
            )
        if error and error != "No images found":
            notes.append(f"Помилка: {error}")

        row_data = [
            row_offset + 1,
            folder_name,
            status,
            len(result.get("sorted_images", [])),
            main_status,
            method_str,
            len(result.get("renamed_files", [])),
            "; ".join(notes),
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = row_fill
            cell.border = _BORDER
            if col in (1, 4, 7):
                cell.alignment = _CELL_ALIGN_CENTER
            else:
                cell.alignment = _CELL_ALIGN_LEFT

    ws.freeze_panes = ws.cell(row=title_row + 1, column=1)
    _apply_col_widths(ws, _SUMMARY_WIDTHS)
    ws.auto_filter.ref = (
        f"A{title_row}:{get_column_letter(len(_SUMMARY_HEADERS))}{title_row}"
    )


# ---------------------------------------------------------------------------
# Аркуш 2 — Деталі файлів
# ---------------------------------------------------------------------------

_DETAIL_HEADERS = [
    "Папка",
    "Оригінальна назва",
    "Нова назва",
    "Категорія",
    "Білий фон (оцінка)",
    "Виявлено текст",
    "Метод аналізу",
    "Помилка",
]

_DETAIL_WIDTHS = [32, 32, 18, 20, 20, 16, 18, 45]


def _build_detail_sheet(ws: Worksheet, results: list[dict]) -> None:
    for col, header in enumerate(_DETAIL_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(_DETAIL_HEADERS))

    detail_row = 2
    alt_fill = _fill(_GRAY)

    for result in results:
        folder_name = result.get("folder_name", "")
        sorted_images = result.get("sorted_images", [])
        renames = result.get("renamed_files", [])

        rename_map = {r["old_name"]: r["new_name"] for r in renames}

        for img in sorted_images:
            old_name = img.get("filename", "")
            new_name = rename_map.get(old_name, "—")
            cat_label = CATEGORY_LABELS_UK.get(img.get("category", ""), img.get("category", ""))
            use_fill = alt_fill if detail_row % 2 == 0 else None

            row_data = [
                folder_name,
                old_name,
                new_name,
                cat_label,
                round(img.get("white_bg_score", 0.0), 3),
                "Так" if img.get("has_text") else "Ні",
                img.get("method", ""),
                img.get("error") or "",
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=detail_row, column=col, value=value)
                cell.border = _BORDER
                cell.alignment = _CELL_ALIGN_LEFT
                if use_fill:
                    cell.fill = use_fill
                if col == 5:
                    cell.alignment = _CELL_ALIGN_CENTER

            detail_row += 1

    ws.freeze_panes = ws.cell(row=2, column=1)
    _apply_col_widths(ws, _DETAIL_WIDTHS)
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_DETAIL_HEADERS))}1"
    )


# ---------------------------------------------------------------------------
# Публічна функція
# ---------------------------------------------------------------------------

def generate_report(results: list[dict], output_path: str) -> str:
    """
    Генерує Excel-звіт та зберігає його у `output_path`.

    Параметри:
      results     — список словників, повернутих process_folder()
      output_path — шлях до файлу .xlsx (якщо файл існує, буде створено новий з міткою часу)

    Повертає абсолютний шлях до збереженого файлу.
    """
    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "Зведений звіт"
    _build_summary_sheet(ws_summary, results)

    ws_detail = wb.create_sheet(title="Деталі файлів")
    _build_detail_sheet(ws_detail, results)

    # Встановлюємо активний аркуш на перший
    wb.active = ws_summary

    # Якщо файл вже існує, створюємо новий з міткою часу
    final_path = output_path
    if os.path.exists(output_path):
        base_dir = os.path.dirname(os.path.abspath(output_path))
        base_name = os.path.basename(output_path)
        name_without_ext, ext = os.path.splitext(base_name)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = f"{name_without_ext}_{timestamp}{ext}"
        final_path = os.path.join(base_dir, new_name)
        logger.info("Файл %s вже існує. Створюємо новий файл: %s", output_path, final_path)

    os.makedirs(os.path.dirname(os.path.abspath(final_path)), exist_ok=True)
    wb.save(final_path)
    logger.info("Звіт збережено: %s", final_path)
    return os.path.abspath(final_path)
